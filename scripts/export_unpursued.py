"""
Export the matched jobs we cannot pursue automatically, for manual chasing.

A job lands here when it matched a candidate well but its description is too
thin for the resume generator to work with, and enrichment could not recover
it. Almost all are Adzuna-sourced: Adzuna truncates at ~500 characters and its
own links are login-gated, so there is nothing to scrape.

The output is a handover sheet: who it was matched for, why it scored, where to
find it, and why automation gave up.

    python -m scripts.export_unpursued
"""
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import db

MIN_GOOD = 1500
OUT = os.path.join(os.path.dirname(__file__), "..",
                   f"unpursued_jobs_{date.today():%Y%m%d}.xlsx")

DEAD_DOMAINS = ("adzuna.com", "youtube.com", "facebook.com", "twitter.com")

HEADERS = [
    ("Candidate", 24), ("Base Resume", 30), ("Score", 7), ("Band", 14),
    ("Job Title", 42), ("Company", 26), ("Location", 26), ("Posted", 11),
    ("Why not pursued", 30), ("Best Link", 52), ("Adzuna Link", 46),
    ("Source", 15), ("Description we have", 60),
]


def reason(row) -> str:
    url = (row["source_url"] or row["apply_url"] or row["job_url"] or "").lower()
    if not url:
        return "No link found"
    if any(d in url for d in DEAD_DOMAINS):
        return "Link is gated/junk - needs manual search"
    return "Description too short, page not scrapeable"


def main():
    with db.get_conn() as conn:
        rows = [dict(r) for r in conn.execute(f"""
            SELECT p.candidate_name, p.base_resume_name, m.score, m.band, m.reason,
                   j.title, j.company_name, j.location, j.posted_date,
                   j.job_url, j.source_url, j.apply_url, j.source,
                   coalesce(j.description,'') description
            FROM resume_job_matches m
            JOIN resume_profiles p ON p.id = m.resume_profile_id
            JOIN keyword_jobs   j ON j.id = m.keyword_job_id
            WHERE p.is_test_account = 0
              AND m.score >= 90
              AND length(coalesce(j.description,'')) < {MIN_GOOD}
            ORDER BY p.candidate_name, m.score DESC
        """).fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Manual chase"

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    for i, (name, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    for r in rows:
        best = r["source_url"] or r["apply_url"] or r["job_url"]
        ws.append([
            r["candidate_name"], r["base_resume_name"], r["score"], r["band"],
            r["title"], r["company_name"], r["location"], str(r["posted_date"] or ""),
            reason(r), best, r["job_url"], r["source"],
            (r["description"] or "")[:400],
        ])

    for row in ws.iter_rows(min_row=2):
        row[12].alignment = Alignment(wrap_text=True, vertical="top")
        for link_cell in (row[9], row[10]):
            if link_cell.value:
                link_cell.hyperlink = link_cell.value
                link_cell.font = Font(color="0563C1", underline="single")

    # summary sheet so whoever picks this up sees the shape of the work
    s = wb.create_sheet("Summary")
    s["A1"], s["A1"].font = "Jobs that need manual chasing", Font(bold=True, size=13)
    s["A3"] = "These matched a candidate at score 90+ but the job description was too"
    s["A4"] = "short for automated resume tailoring, and could not be recovered."
    s["A5"] = "Adzuna truncates descriptions at ~500 chars and gates its own links."

    per_cand: dict[str, int] = {}
    per_reason: dict[str, int] = {}
    for r in rows:
        per_cand[r["candidate_name"]] = per_cand.get(r["candidate_name"], 0) + 1
        per_reason[reason(r)] = per_reason.get(reason(r), 0) + 1

    s["A7"], s["A7"].font = "Total", Font(bold=True)
    s["B7"] = len(rows)
    s["A9"], s["A9"].font = "By candidate", Font(bold=True)
    for i, (k, v) in enumerate(sorted(per_cand.items(), key=lambda x: -x[1]), start=10):
        s[f"A{i}"], s[f"B{i}"] = k, v
    start = 10 + len(per_cand) + 1
    s[f"A{start}"], s[f"A{start}"].font = "By reason", Font(bold=True)
    for i, (k, v) in enumerate(sorted(per_reason.items(), key=lambda x: -x[1]), start=start + 1):
        s[f"A{i}"], s[f"B{i}"] = k, v
    s.column_dimensions["A"].width = 46
    s.column_dimensions["B"].width = 10

    wb.save(OUT)
    print(f"{len(rows)} unpursued matches -> {os.path.abspath(OUT)}")
    print("by candidate:", dict(sorted(per_cand.items(), key=lambda x: -x[1])))
    print("by reason:", per_reason)


if __name__ == "__main__":
    main()
