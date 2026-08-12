"""
The manual-chase handover sheet.

A job lands here when it matched a candidate well but its description is too
thin for the resume generator, and enrichment could not recover it. Almost all
are Adzuna-sourced: Adzuna truncates at ~500 characters and gates its own links,
so there is nothing left to scrape.

These are NOT bad matches — they are good matches we cannot automate. Somebody
should chase them by hand, which is what this workbook is for.

This is a library, not a script, so all three callers produce an identical file:

    app/manual_chase_tab.py    download button in the UI
    scripts/export_unpursued   command line
    scripts/daily_cycle        written to data/exports/ every night

Returns bytes rather than writing to disk, because Streamlit needs bytes and a
file on the server is no use to someone using the app over the tunnel.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import db
from app.quality import MIN_DESCRIPTION, best_link, blocked_reason

HEADERS = [
    ("Candidate", 24), ("Base Resume", 30), ("Score", 7), ("Band", 14),
    ("Job Title", 42), ("Company", 26), ("Location", 26), ("Posted", 11),
    ("Why not pursued", 32), ("Best Link", 52), ("Aggregator Link", 46),
    ("Source", 15), ("Description we have", 60),
]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")


def unpursued_rows(min_score: int = 90,
                   d_from: date | None = None,
                   d_to: date | None = None,
                   candidate: str | None = None) -> list[dict]:
    """
    Matches good enough to pursue but too thin to automate.

    Date and candidate filters mirror the Review tab, so what an operator sees
    on screen is exactly what they download.
    """
    sql = f"""
        SELECT p.candidate_name, p.base_resume_name, m.score, m.band, m.reason,
               j.title, j.company_name, j.location, j.posted_date,
               j.job_url, j.source_url, j.apply_url, j.source,
               coalesce(j.description,'') description
        FROM resume_job_matches m
        JOIN resume_profiles p ON p.id = m.resume_profile_id
        JOIN keyword_jobs   j ON j.id = m.keyword_job_id
        WHERE p.is_test_account = 0
          AND m.score >= ?
          AND length(coalesce(j.description,'')) < {MIN_DESCRIPTION}
    """
    params: list = [min_score]
    if d_from and d_to:
        sql += " AND j.posted_date BETWEEN ? AND ?"
        params += [d_from.isoformat(), d_to.isoformat()]
    if candidate and candidate != "All":
        sql += " AND p.candidate_name = ?"
        params.append(candidate)
    sql += " ORDER BY p.candidate_name, m.score DESC"

    with db.get_conn() as conn:
        return [dict(r) for r in conn.execute(sql, params).fetchall()]


def build_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Manual chase"

    for i, (name, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    for r in rows:
        ws.append([
            r["candidate_name"], r["base_resume_name"], r["score"], r["band"],
            r["title"], r["company_name"], r["location"], str(r["posted_date"] or ""),
            blocked_reason(r) or "", best_link(r), r["job_url"], r["source"],
            (r["description"] or "")[:400],
        ])

    for row in ws.iter_rows(min_row=2):
        row[12].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in (row[9], row[10]):
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = LINK_FONT

    _summary_sheet(wb, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _summary_sheet(wb: Workbook, rows: list[dict]) -> None:
    """So whoever picks this up sees the shape of the work before the detail."""
    s = wb.create_sheet("Summary")
    s["A1"], s["A1"].font = "Jobs that need manual chasing", Font(bold=True, size=13)
    s["A3"] = "These matched a candidate at a high score but the job description was"
    s["A4"] = "too short for automated resume tailoring, and could not be recovered."
    s["A5"] = "Adzuna truncates descriptions at ~500 chars and gates its own links."

    per_cand: dict[str, int] = {}
    per_reason: dict[str, int] = {}
    for r in rows:
        per_cand[r["candidate_name"]] = per_cand.get(r["candidate_name"], 0) + 1
        why = blocked_reason(r) or "unknown"
        per_reason[why] = per_reason.get(why, 0) + 1

    s["A7"], s["A7"].font = "Total", Font(bold=True)
    s["B7"] = len(rows)

    s["A9"], s["A9"].font = "By candidate", Font(bold=True)
    for i, (k, v) in enumerate(sorted(per_cand.items(), key=lambda x: -x[1]), start=10):
        s[f"A{i}"], s[f"B{i}"] = k, v

    start = 10 + len(per_cand) + 1
    s[f"A{start}"], s[f"A{start}"].font = "By reason", Font(bold=True)
    for i, (k, v) in enumerate(sorted(per_reason.items(), key=lambda x: -x[1]), start=start + 1):
        s[f"A{i}"], s[f"B{i}"] = k, v

    s.column_dimensions["A"].width = 46
    s.column_dimensions["B"].width = 10


def summarise(rows: list[dict]) -> dict:
    """Counts for the UI, computed the same way the workbook computes them."""
    per_reason: dict[str, int] = {}
    per_cand: dict[str, int] = {}
    for r in rows:
        why = blocked_reason(r) or "unknown"
        per_reason[why] = per_reason.get(why, 0) + 1
        per_cand[r["candidate_name"]] = per_cand.get(r["candidate_name"], 0) + 1
    return {"total": len(rows), "by_reason": per_reason, "by_candidate": per_cand}


def filename(d_from: date | None = None, d_to: date | None = None) -> str:
    if d_from and d_to:
        return f"manual_chase_{d_from:%Y%m%d}_{d_to:%Y%m%d}.xlsx"
    return f"manual_chase_{date.today():%Y%m%d}.xlsx"
